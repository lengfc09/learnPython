import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os

wb = load_workbook("CDDA67.xlsx")
sheetnames = wb.get_sheet_names()
ws1 = wb[sheetnames[0]]
tbsheetnames = wb.get_sheet_names()[1:]
ProductName = ws1["A2"].value
NoTable = ws1["B2"].value
tables = {}
for i, sheetname in enumerate(tbsheetnames):
    tables[ws1["C" + str(i + 4)].value] = pd.read_excel("CDDA67.xlsx", sheetname=sheetname, skiprows=[0], index_col=[0])

# print(tables.keys())
# print(tables["预定附加费用率"].head())
# print(tables["恶性肿瘤疾病发生率"].head())
# tableDD = tables["恶性肿瘤疾病发生率"]
# print("3周岁男性恶性肿瘤疾病发生率:", tableDD.loc[3, "男性"])
for key in tables:
    print(key)
    print(tables[key].head(5))
    print("next one")
